#!/usr/bin/env python3
"""
Gene Panel Consolidation & HGNC Symbol Validation (automatic JSON download)

Run this script in a folder containing your source gene-panel .xlsx files.

It will:
1. Download the HGNC complete-set JSON automatically when it is missing.
2. Keep prior HGNC JSON files instead of overwriting them.
3. Save each newly downloaded HGNC file with today's date, for example:
       hgnc_complete_set_2026-08-28.json
4. Reuse the most recent dated HGNC JSON file for up to 30 days.
5. Consolidate panel workbooks and update aliases/old gene names to current
   HGNC-approved symbols.

Requirements:
    pip install pandas openpyxl requests

Usage:
    python3 consolidate_gene_panels_auto_hgnc_json_dated.py

Optional output name:
    python3 consolidate_gene_panels_auto_hgnc_json_dated.py my_output.xlsx
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_FILE = sys.argv[1] if len(sys.argv) > 1 else "output_gene_panel.xlsx"
HGNC_URL = "https://storage.googleapis.com/public-download-files/hgnc/json/json/hgnc_complete_set.json"
HGNC_FILE_PREFIX = "hgnc_complete_set_"
HGNC_FILE_SUFFIX = ".json"
HGNC_MAX_AGE_DAYS = 30
HGNC_DOWNLOAD_TIMEOUT_SECONDS = 90

OUT_COLS = [
    "panel_name",
    "Gene",
    "Disease_Name",
    "Inheritance",
    "Citation_(OMIM or PUBMED)",
    "Relationship_to_phenotype",
    "additional_notes",
]

FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_RED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
FILL_NONE = PatternFill(fill_type=None)


def extract_hgnc_records(payload: Any) -> list[dict[str, Any]]:
    """Extract HGNC records from the standard HGNC complete-set JSON structure."""
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        raise ValueError("HGNC JSON root is neither an object nor a list.")

    response = payload.get("response")
    if isinstance(response, dict) and isinstance(response.get("docs"), list):
        return [item for item in response["docs"] if isinstance(item, dict)]

    for key in ("docs", "records", "results", "data"):
        candidate = payload.get(key)
        if isinstance(candidate, list):
            return [item for item in candidate if isinstance(item, dict)]
        if isinstance(candidate, dict):
            for nested_key in ("docs", "records", "results"):
                nested = candidate.get(nested_key)
                if isinstance(nested, list):
                    return [item for item in nested if isinstance(item, dict)]

    raise ValueError("Could not find a list of gene records in the HGNC JSON file.")


def validate_hgnc_json(path: Path) -> int:
    """Return the record count if the JSON is a usable HGNC complete-set file."""
    with path.open("r", encoding="utf-8") as handle:
        records = extract_hgnc_records(json.load(handle))
    if not records:
        raise ValueError("No records found in HGNC JSON.")
    if not any(str(record.get("symbol", "")).strip() for record in records):
        raise ValueError("HGNC JSON has no usable 'symbol' fields.")
    return len(records)


def dated_hgnc_filename() -> str:
    return f"{HGNC_FILE_PREFIX}{date.today().isoformat()}{HGNC_FILE_SUFFIX}"


def find_most_recent_hgnc_file(folder: Path) -> Path | None:
    """Find the newest valid date-stamped HGNC JSON file in the current folder."""
    candidates = sorted(
        folder.glob(f"{HGNC_FILE_PREFIX}????-??-??{HGNC_FILE_SUFFIX}"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for candidate in candidates:
        try:
            record_count = validate_hgnc_json(candidate)
            print(f"[INFO] Found valid HGNC JSON: {candidate.name} ({record_count:,} records)")
            return candidate
        except Exception as exc:
            print(f"[WARNING] Ignoring invalid HGNC JSON '{candidate.name}': {exc}")
    return None


def ensure_hgnc_complete_set(folder: Path) -> Path | None:
    """Reuse a recent dated JSON file or download a new dated copy without overwriting old ones."""
    existing = find_most_recent_hgnc_file(folder)
    if existing is not None:
        age_days = (time.time() - existing.stat().st_mtime) / 86400
        if age_days < HGNC_MAX_AGE_DAYS:
            print(f"[INFO] Using {existing.name}; it is {age_days:.1f} days old.")
            return existing
        print(f"[INFO] Latest HGNC JSON is {age_days:.1f} days old; downloading a fresh dated copy.")
    else:
        print("[INFO] No usable local HGNC JSON found; downloading a fresh dated copy.")

    output_path = folder / dated_hgnc_filename()
    if output_path.exists():
        try:
            record_count = validate_hgnc_json(output_path)
            print(f"[INFO] Today's HGNC file already exists: {output_path.name} ({record_count:,} records)")
            return output_path
        except Exception:
            print(f"[WARNING] Today's HGNC file is invalid and will be replaced: {output_path.name}")

    temporary_path = output_path.with_name(output_path.name + ".download")
    try:
        print("[INFO] Downloading current HGNC complete-set JSON...")
        response = requests.get(HGNC_URL, stream=True, timeout=HGNC_DOWNLOAD_TIMEOUT_SECONDS)
        response.raise_for_status()

        bytes_written = 0
        with temporary_path.open("wb") as handle:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    handle.write(chunk)
                    bytes_written += len(chunk)

        if bytes_written == 0:
            raise RuntimeError("Downloaded HGNC JSON is empty.")

        record_count = validate_hgnc_json(temporary_path)
        os.replace(temporary_path, output_path)
        print(
            f"[INFO] HGNC JSON saved: {output_path.name} "
            f"({bytes_written / 1024 / 1024:.1f} MB; {record_count:,} records)"
        )
        return output_path

    except Exception as exc:
        temporary_path.unlink(missing_ok=True)
        print(f"[WARNING] Could not download or validate HGNC JSON: {exc}")
        if existing is not None:
            print(f"[WARNING] Continuing with older local HGNC JSON: {existing.name}")
            return existing
        print("[WARNING] No usable HGNC reference is available; symbols will not be validated.")
        return None


def field_values(record: dict[str, Any], field: str) -> list[str]:
    """Read HGNC alias/previous-symbol values whether stored as lists or text."""
    value = record.get(field, [])
    if value is None:
        return []
    raw_values = value if isinstance(value, list) else re.split(r"[|,;]+", str(value))
    return [str(item).strip().strip('"') for item in raw_values if str(item).strip()]


def build_hgnc_lookup(hgnc_path: Path) -> dict[str, str]:
    """Map current HGNC symbols, aliases, and previous symbols to approved symbols."""
    try:
        with hgnc_path.open("r", encoding="utf-8") as handle:
            records = extract_hgnc_records(json.load(handle))
    except Exception as exc:
        print(f"[WARNING] Could not read HGNC JSON '{hgnc_path.name}': {exc}")
        return {}

    lookup: dict[str, str] = {}
    for record in records:
        approved = str(record.get("symbol", "")).strip()
        if approved:
            lookup[approved.upper()] = approved

    for record in records:
        approved = str(record.get("symbol", "")).strip()
        if not approved:
            continue
        for field in ("alias_symbol", "prev_symbol"):
            for symbol in field_values(record, field):
                lookup.setdefault(symbol.upper(), approved)

    print(f"[INFO] HGNC lookup built: {len(lookup):,} entries (approved, alias, previous symbols)")
    return lookup


def parse_panel_file(path: Path, panel_name: str) -> list[dict[str, str]]:
    """Read all sheets with an exact Gene column and normalize their supported fields."""
    rows: list[dict[str, str]] = []
    try:
        excel = pd.ExcelFile(path)
    except Exception as exc:
        print(f"[WARNING] Could not open {path.name}: {exc}")
        return rows

    for sheet_name in excel.sheet_names:
        try:
            dataframe = excel.parse(sheet_name, dtype=str).dropna(how="all").reset_index(drop=True)
        except Exception as exc:
            print(f"[WARNING] Could not parse sheet '{sheet_name}' in {path.name}: {exc}")
            continue

        if dataframe.empty:
            continue
        if "Gene" not in dataframe.columns:
            print(f"[WARNING] No exact 'Gene' column in sheet '{sheet_name}' of {path.name}; skipping sheet.")
            continue

        for _, source_row in dataframe.iterrows():
            gene = str(source_row.get("Gene", "")).strip()
            if not gene or gene.lower() in {"nan", "gene"}:
                continue

            output_row = {column: "" for column in OUT_COLS}
            output_row["panel_name"] = panel_name
            for column in OUT_COLS:
                if column != "panel_name" and column in dataframe.columns:
                    value = str(source_row.get(column, "")).strip()
                    output_row[column] = "" if value.lower() == "nan" else value
            rows.append(output_row)

    return rows


def is_input_workbook(path: Path, output_path: Path) -> bool:
    lower_name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and path.resolve() != output_path.resolve()
        and not lower_name.startswith("~$")
        and not lower_name.endswith("_omim.xlsx")
    )


def add_legend(workbook: Workbook) -> None:
    legend = workbook.create_sheet("Legend")
    legend.append(["Color", "Meaning"])
    legend.append(["White (no fill)", "Current approved HGNC symbol — no change"])
    legend.append(["Yellow", "Alias or previous symbol — updated to current approved HGNC symbol"])
    legend.append(["Red", "Symbol not found in HGNC — requires manual review"])
    for cell in legend[1]:
        cell.font = Font(bold=True)
    legend["A3"].fill = FILL_YELLOW
    legend["A4"].fill = FILL_RED
    legend.column_dimensions["A"].width = 20
    legend.column_dimensions["B"].width = 90


def main() -> None:
    current_folder = Path.cwd()
    output_path = (current_folder / OUTPUT_FILE).resolve()

    hgnc_path = ensure_hgnc_complete_set(current_folder)
    hgnc_lookup = build_hgnc_lookup(hgnc_path) if hgnc_path else {}

    input_files = sorted(
        path for path in current_folder.iterdir()
        if path.is_file() and is_input_workbook(path, output_path)
    )
    if not input_files:
        print("[ERROR] No eligible .xlsx panel files found in the current folder.")
        print("[INFO] The output workbook, temporary Excel files, and *_omim.xlsx files are ignored.")
        raise SystemExit(1)

    print(f"[INFO] Found {len(input_files)} panel file(s):")
    for path in input_files:
        print(f"  {path.name}")

    all_rows: list[dict[str, str]] = []
    for path in input_files:
        rows = parse_panel_file(path, path.stem)
        print(f"[INFO] {path.name:60s} -> {len(rows):4d} gene row(s)")
        all_rows.extend(rows)

    if not all_rows:
        print("[ERROR] No gene rows were parsed. Input sheets need an exact header named 'Gene'.")
        raise SystemExit(1)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Gene_Panels"
    worksheet.append(OUT_COLS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    gene_column = OUT_COLS.index("Gene") + 1
    stats = {"unchanged": 0, "renamed": 0, "not_found": 0, "unvalidated": 0}

    for row in all_rows:
        original = row["Gene"].strip()
        resolved = original
        fill = FILL_NONE

        if hgnc_lookup:
            approved = hgnc_lookup.get(original.upper())
            if approved is None:
                fill = FILL_RED
                stats["not_found"] += 1
                print(f"  [NOT FOUND] '{original}'")
            elif approved == original:
                stats["unchanged"] += 1
            else:
                resolved = approved
                fill = FILL_YELLOW
                stats["renamed"] += 1
                print(f"  [RENAMED] '{original}' -> '{approved}'")
        else:
            stats["unvalidated"] += 1

        row["Gene"] = resolved
        worksheet.append([row[column] for column in OUT_COLS])
        worksheet.cell(row=worksheet.max_row, column=gene_column).fill = fill

    for column_index, column_name in enumerate(OUT_COLS, start=1):
        maximum = len(column_name)
        for row_cells in worksheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            value = row_cells[0].value
            if value is not None:
                maximum = max(maximum, min(len(str(value)), 80))
        worksheet.column_dimensions[get_column_letter(column_index)].width = maximum + 2

    worksheet.freeze_panes = "A2"
    add_legend(workbook)
    workbook.save(output_path)

    print("\n" + "=" * 60)
    print(f"Output: {output_path}")
    print(f"Total genes: {len(all_rows)}")
    if hgnc_lookup:
        print(f"Unchanged approved symbols: {stats['unchanged']}")
        print(f"Renamed symbols (yellow): {stats['renamed']}")
        print(f"Not found (red): {stats['not_found']}")
    else:
        print(f"Not HGNC-validated: {stats['unvalidated']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
