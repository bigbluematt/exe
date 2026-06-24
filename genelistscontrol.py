#!/usr/bin/env python3
"""
Compare master panel Excel (excel2) against all `_omim` Excel files (excel1s),
panel by panel, and write an output Excel where each column is a panel_name.

Requirements:
- pandas
- openpyxl

Usage:
- Put this script in the same directory as MasterFinalGeneList_MN.xlsx.
- Run: python compare_panels.py
- It will walk the current directory and all subdirectories, find all files
  containing `_omim` in the filename, and produce output/panel_gene_diff_by_panel.xlsx.
"""

import os
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

EXCEL2_FILENAME = "MasterFinalGeneList_MN.xlsx"   # master file (excel2)
PANEL_COL_NAMES = ["panel_name", "panelname"]     # allowed panel column names (case-insensitive)
GENE_COL_NAME = "Gene"                           # expected gene column name

OUTPUT_DIR = "output"
OUTPUT_FILENAME = "panel_gene_diff_by_panel.xlsx"

# ---------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------

def find_panel_column(df):
    """Find the panel_name column in a DataFrame."""
    for c in df.columns:
        if c.lower() in PANEL_COL_NAMES:
            return c
    # Fallback: first column
    return df.columns[0]


def find_gene_column(df):
    """Find the gene column in a DataFrame."""
    # Prefer exact 'Gene' if present
    if GENE_COL_NAME in df.columns:
        return GENE_COL_NAME
    # Otherwise any column containing 'gene' in its name
    for c in df.columns:
        if "gene" in c.lower():
            return c
    # Fallback: second column
    return df.columns[1]


def normalize_str(x):
    """Normalize panel_name / gene strings safely."""
    if pd.isna(x):
        return None
    return str(x).strip()


# ---------------------------------------------------------------------
# Load excel2 (master) and build G2(p)
# ---------------------------------------------------------------------

if not os.path.exists(EXCEL2_FILENAME):
    raise FileNotFoundError(f"Master file '{EXCEL2_FILENAME}' not found in current directory.")

excel2 = pd.read_excel(EXCEL2_FILENAME)
panel_col2 = find_panel_column(excel2)
gene_col2 = find_gene_column(excel2)

excel2_panel_genes = defaultdict(set)  # panel_name -> set of genes (from excel2)

for _, row in excel2.iterrows():
    p = normalize_str(row[panel_col2])
    g = normalize_str(row[gene_col2])
    if p and g:
        excel2_panel_genes[p].add(g)

P2 = set(excel2_panel_genes.keys())

# ---------------------------------------------------------------------
# Walk directory tree and collect all excel1 files (containing '_omim')
# ---------------------------------------------------------------------

excel1_paths = []

for root, dirs, files in os.walk("."):
    for f in files:
        if "_omim" in f and f.lower().endswith((".xlsx", ".xls")):
            # Exclude the master file if its name happens to contain '_omim'
            if os.path.basename(f) == EXCEL2_FILENAME:
                continue
            excel1_paths.append(os.path.join(root, f))

# ---------------------------------------------------------------------
# Build G1(p): union of genes for each panel across all excel1s
# ---------------------------------------------------------------------

excel1_panel_genes = defaultdict(set)  # panel_name -> set of genes (from all excel1s)

for path in excel1_paths:
    try:
        df = pd.read_excel(path)
    except Exception as e:
        print(f"Warning: could not read '{path}': {e}")
        continue

    panel_col1 = find_panel_column(df)
    gene_col1 = find_gene_column(df)

    for _, row in df.iterrows():
        p = normalize_str(row[panel_col1])
        g = normalize_str(row[gene_col1])
        if p and g:
            excel1_panel_genes[p].add(g)

P1 = set(excel1_panel_genes.keys())

# ---------------------------------------------------------------------
# Panel presence comparison
# ---------------------------------------------------------------------

panels_only_in_excel2 = sorted(P2 - P1)  # exist in master only
panels_only_in_excel1 = sorted(P1 - P2)  # exist in excel1s only
all_panels = sorted(P1.union(P2))        # union of all panels

# ---------------------------------------------------------------------
# Build per-panel gene differences
# G2(p) and G1(p) may be empty for some panels (e.g., panel only on one side)
# ---------------------------------------------------------------------

panel_diffs = {}  # panel -> dict with 'missing_in_excel1' and 'added_in_excel1'

for p in all_panels:
    G2 = excel2_panel_genes.get(p, set())
    G1 = excel1_panel_genes.get(p, set())

    missing_in_excel1 = sorted(G2 - G1)  # present in excel2, absent from excel1s
    added_in_excel1 = sorted(G1 - G2)    # present in excel1s, absent from excel2

    panel_diffs[p] = {
        "missing_in_excel1": missing_in_excel1,
        "added_in_excel1": added_in_excel1,
    }

# ---------------------------------------------------------------------
# Create output workbook
# ---------------------------------------------------------------------

os.makedirs(OUTPUT_DIR, exist_ok=True)
wb = Workbook()

# Main sheet: per-panel gene differences
ws_main = wb.active
ws_main.title = "panel_gene_diff"

# Colors (Excel ARGB)
green_fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")  # light green
red_fill = PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid")    # light red

col_idx = 1

for panel_name in all_panels:
    # Header = panel_name
    ws_main.cell(row=1, column=col_idx, value=panel_name)

    row_idx = 2
    diff = panel_diffs[panel_name]

    # First: genes present in excel2 but missing in excel1s (green)
    for gene in diff["missing_in_excel1"]:
        cell = ws_main.cell(row=row_idx, column=col_idx, value=gene)
        cell.fill = green_fill
        row_idx += 1

    # Then: genes present in excel1s but missing in excel2 (red)
    for gene in diff["added_in_excel1"]:
        cell = ws_main.cell(row=row_idx, column=col_idx, value=gene)
        cell.fill = red_fill
        row_idx += 1

    col_idx += 1

# Second sheet: panel presence summary
ws_panels = wb.create_sheet(title="panel_presence")

ws_panels.cell(row=1, column=1, value="Panels only in excel2")
ws_panels.cell(row=1, column=2, value="Panels only in excel1s")

max_rows = max(len(panels_only_in_excel2), len(panels_only_in_excel1))
for i in range(max_rows):
    if i < len(panels_only_in_excel2):
        ws_panels.cell(row=2 + i, column=1, value=panels_only_in_excel2[i])
    if i < len(panels_only_in_excel1):
        ws_panels.cell(row=2 + i, column=2, value=panels_only_in_excel1[i])

# Save workbook
output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILENAME)
wb.save(output_path)

print(f"Written: {output_path}")